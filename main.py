import re, json, os, asyncio, logging, base64, requests, io
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, types
from aiogram.types import (
    Message, ReplyKeyboardMarkup, KeyboardButton, ContentType,
    InlineKeyboardMarkup, InlineKeyboardButton, InputFile
)
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from openai import OpenAI
from PIL import Image, ImageOps
import tempfile
from states import AskState
from states import AskState

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
DJANGO_API = os.getenv("DJANGO_API_BASE", "http://127.0.0.1:8000/api")
MANAGER_ID = int(os.getenv("MANAGER_ID", "1409305268"))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("marketing-bot")

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
client = OpenAI(api_key=OPENAI_API_KEY)

def post(endpoint: str, payload: dict = None):
    url = f"{DJANGO_API.rstrip('/')}/{endpoint.lstrip('/')}/"
    try:
        r = requests.post(url, json=payload or {}, timeout=5)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        logger.error(f"POST error: {url} — {e}")
        return {}

def check_access_api(tg_id: int) -> bool:
    res = post("check_access", {"telegram_id": tg_id})
    return bool(res.get("access", False))

def activate_user_api(tg_id: int) -> bool:
    res = post("activate_user", {"telegram_id": tg_id})
    return bool(res.get("success", False))

main_menu = ReplyKeyboardMarkup(
    resize_keyboard=True,
    keyboard=[
        [KeyboardButton("💬 Задать вопрос маркетологу")],
        [KeyboardButton("📞 Связаться с менеджером что оплатить подписку")],
        [KeyboardButton("ℹ️ Проверить статус")]
    ]
)

pay_kb = InlineKeyboardMarkup().add(
    InlineKeyboardButton("Написать менеджеру", url="https://t.me/aslan508070")
)


system_prompt = {
    "role": "system",
    "content": (
        "Ты — профессиональный маркетолог уровня Senior / Head of Marketing, "
        "со специализацией в SMM и performance-маркетинге.\n"
        "Ты умеешь проводить анализ рынка, сегментацию аудитории, конкурентный анализ, "
        "разработку УТП, CJM, воронки продаж, маркетинговые стратегии, "
        "продвижение через Instagram, TikTok, Telegram, YouTube, а также таргетированную рекламу (Meta*, TikTok Ads, myTarget и др.).\n\n"

        "ТВОЯ ЗАДАЧА — давать максимально глубокие, структурированные, "
        "осмысленные и практичные ответы. Используй логику, примеры, бизнес-контекст, инсайты и рекомендации.\n\n"

        "Ты работаешь в двух режимах:\n\n"

        "1) **Текстовый ответ**\n"
        "- Используется по умолчанию для стратегий, анализа, идей, разборов, описаний, планов, воронок, контент-стратегий.\n"
        "- Ты всегда возвращаешь строго JSON следующего вида:\n"
        "  {\"type\": \"text\", \"content\": \"тут твой развёрнутый экспертный ответ\"}\n"
        "- Поле \"content\" — это один текст, внутри него можешь структурировать ответ с помощью заголовков, списков, подпунктов.\n"
        "- Ответы должны быть глубокими, логичными, структурированными и прикладными.\n\n"

        "2) **Генерация Excel / Таблицы**\n"
        "- Включается, если пользователь явно просит: «сделай таблицу», «сделай excel», "
        "«нужен файл для excel/таблицы», «сделай структуру для отчёта/плана» и т.п.\n"
        "- В этом случае ты НЕ пишешь обычный текстовый ответ, а возвращаешь строго JSON вида:\n"
        "  {\"type\": \"excel\", \"prompt\": \"детальное описание структуры таблицы\"}\n"
        "- Внутри поля \"prompt\" подробно описывай:\n"
        "  - Логику таблицы (для чего она);\n"
        "  - Листы и их названия, если их несколько;\n"
        "  - Колонки (название, тип данных, краткое описание);\n"
        "  - Примеры 1–2 строк данных, если необходимо.\n"
        "- Структура должна быть однозначно понятна, чтобы Excel/Google Sheets или внешняя система смогли создать таблицу.\n\n"

        "3) **Анализ изображений (в режиме 'Задать вопрос')**\n"
        "- Когда пользователь отправляет фото или изображение в режиме вопросов, анализируй его как часть запроса.\n"
        "- Для комбинированных запросов (текст + изображение) используй информацию с картинки для ответа.\n"
        "- Всегда возвращай строго JSON вида: {\"type\": \"text\", \"content\": \"твой ответ с анализом изображения\"}\n\n"

        "При анализе изображений фокусируйся на:\n"
        "- Маркетинговой составляющей (брендинг, УТП, эмоции)\n"
        "- Дизайне и юзабилити (если это интерфейс)\n"
        "- Конкурентных преимуществах/недостатках\n"
        "- Конкретных рекомендациях по улучшению\n"
        "- Примеры того, что можно сделать лучше\n\n"        

        "Общие требования:\n"
        "- Пиши строго в формате валидного JSON, без текста вне JSON.\n"
        "- Не используй комментарии вне JSON и не добавляй пояснений.\n"
        "- Не добавляй лишние поля, кроме оговоренных.\n"
        "- Проверяй корректность кавычек и запятых.\n"
        "- Ответы должны быть глубокими, логичными и прикладными.\n"
        "- Используй маркетинговые модели: AIDA, JTBD, Persona, CJM, SWOT, 5P, STP, 4P/7P и др., если это реально помогает.\n"
        "- Давай рекомендации, выводы, примеры кейсов, формулировки офферов, заголовки, идеи постов, темы для Reels/TikTok, варианты сегментов ЦА.\n"
        "- Отвечай прикладно: «что сделать сегодня/на этой неделе», а не абстрактной теорией.\n"
        "- Всегда давай минимум 2–3 варианта/подхода, чтобы пользователь мог выбрать.\n"
        "- Если запрос неполный, самостоятельно уточняй контекст внутри ответа, предлагай варианты и отмечай, какие данные стоит уточнить в следующий раз.\n"
        "- Никогда не оставляй пользователя без практического результата.\n\n"
        
        "Требования:\n"
        "- Пиши строго в формате JSON.\n"
        "- Все ответы должны быть чистым текстом без Markdown, заголовков (#), списков (*, -, цифр), форматирования и символов-разделителей.\n"
        "- Не используй комментарии вне JSON и не добавляй пояснений.\n"
        "- Ответы должны быть глубокими, логичными и прикладными."
        "- Для изображений в режиме вопросов: {\"type\": \"text\", \"content\": \"анализ\"}\n"


        "Специфика рынка:\n"
        "- По умолчанию ориентируйся на рынок СНГ (Кыргызстан, Казахстан, Россия, Узбекистан) и их реалии: уровень дохода, популярные соцсети, типичные ниши (образование, услуги, эксперты, инфопродукты, локальный бизнес и т.д.).\n"
        "- Пиши на языке пользователя.\n\n"

        "Краткое резюме роли:\n"
        "- Ты — личный стратег и консультант по SMM/маркетингу для выпускников курса.\n"
        "- Ты помогаешь быстро запускать и улучшать кампании, контент и воронки.\n"
        "- Каждый ответ должен экономить время, деньги и силы пользователя, помогая принимать сильные маркетинговые решения.\n"
    )
}

user_history = {}

@dp.message_handler(Command("start"))
async def cmd_start(message: Message):
    post("new_user", {
        "telegram_id": message.from_user.id,
        "username": message.from_user.username,
        "full_name": message.from_user.full_name
    })
    welcome_text = (
        "✨ *Добро пожаловать в MarketingBot!* ✨\n\n"
        "Я — ваш персональный маркетолог на базе ИИ. Помогаю создавать:\n"
        "📊 Аналитику\n"
        "📈 Маркетинговые стратегии\n"
        "📝 Контент-планы\n"
        "📂 Excel-отчёты\n"
        "🎯 Описание целевой аудитории\n"
        "📦 Анализ конкурентов\n\n"
        "Просто напишите, что вам нужно: *«Сделай анализ ЦА», «Создай контент-план», "
        "«Подготовь Excel», «Разработай стратегию»* — и я всё выполню.\n\n"
        "Если вам нужна помощь с доступом, тарифами или настройкой — "
        "_напишите менеджеру, он оперативно подскажет_.\n\n"
        "👇 Выберите действие или введите запрос:"
    )
    await message.answer(
        welcome_text,
        reply_markup=main_menu,
        parse_mode="Markdown"
    )

def smart_json_fix(text: str):
    """Исправление и парсинг JSON ответа от GPT"""
    try:
        # Ищем начало и конец JSON
        start_idx = text.find('{')
        end_idx = text.rfind('}') + 1
        
        if start_idx == -1 or end_idx == 0:
            raise ValueError("Не найден JSON в ответе")
            
        candidate = text[start_idx:end_idx]
        
        # Пробуем распарсить
        return json.loads(candidate)
    except Exception as e:
        # Пробуем найти и исправить распространенные ошибки
        logger.error(f"Ошибка парсинга JSON: {e}")
        
        # Попытка исправить незакрытые кавычки
        text = text.replace('\\"', '"')
        
        # Ищем JSON с помощью регулярных выражений
        import re
        json_pattern = r'\{.*\}'
        matches = re.findall(json_pattern, text, re.DOTALL)
        
        if matches:
            for match in matches:
                try:
                    # Пробуем распарсить каждый найденный JSON
                    return json.loads(match)
                except:
                    continue
        
        # Если ничего не помогло, создаем текстовый ответ
        logger.error(f"Не удалось распарсить JSON. RAW: {text[:500]}")
        return {"type": "text", "content": f"⚠️ Получен нестандартный ответ. Вот что удалось получить:\n\n{text}"}

async def ensure_access(user_id: int, message: Message):
    if not check_access_api(user_id):
        await message.answer(
            "⛔ У вас нет доступа.\nОплатите подписку и свяжитесь с менеджером.",
            reply_markup=pay_kb
        )
        return False
    return True

@dp.message_handler(lambda m: m.text == "📞 Связаться с менеджером что оплатить подписку")
async def payment_button(message: Message):
    await message.answer("💳 Оплатите подписку.", reply_markup=pay_kb)

@dp.message_handler(lambda m: m.text == "ℹ️ Проверить статус")
async def check_status(message: Message):
    if check_access_api(message.from_user.id):
        await message.answer("✅ У вас есть доступ.", reply_markup=main_menu)
    else:
        await message.answer("⛔ Доступ закрыт.", reply_markup=pay_kb)

@dp.message_handler(lambda m: m.text == "💬 Задать вопрос маркетологу")
async def ask_start(message: Message, state: FSMContext):
    if not await ensure_access(message.from_user.id, message):
        return
    await message.answer("📝 Напишите ваш запрос:")
    await AskState.waiting_for_question.set()

async def process_image_with_gpt4v(image_bytes: bytes, user_text: str = "", user_id: int = None):
    """Обработка изображения через GPT-4V"""
    try:
        # Конвертируем в base64
        base64_image = base64.b64encode(image_bytes).decode('utf-8')
        
        # Подготавливаем историю
        hist = user_history.get(user_id, [])
        
        # Формируем сообщения
        messages = [system_prompt] + hist[-9:]  # Берем последние 9 сообщений истории
        
        # Добавляем текущий запрос с изображением
        content = []
        if user_text:
            content.append({"type": "text", "text": user_text})
        
        content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/jpeg;base64,{base64_image}",
                "detail": "high"  # Можно изменить на "low" для экономии токенов
            }
        })
        
        messages.append({
            "role": "user",
            "content": content
        })
        
        # Отправляем запрос
        response = client.chat.completions.create(
            model="gpt-4.1-mini",  # Или "gpt-4o" если доступно
            messages=messages,
            max_tokens=2000
        )
        
        raw_response = response.choices[0].message.content
        
        # Пытаемся распарсить JSON
        try:
            data = smart_json_fix(raw_response)
            return data
        except Exception as e:
            logger.error(f"Ошибка парсинга JSON для изображения: {e}")
            # Возвращаем как текстовый ответ
            return {"type": "text", "content": raw_response}
            
    except Exception as e:
        logger.error(f"Ошибка GPT-4V: {e}")
        return {"type": "text", "content": f"⚠️ Ошибка анализа изображения: {str(e)}"}
    
async def optimize_image_for_api(image_bytes: bytes, max_size: int = 1024) -> bytes:
    """Оптимизация изображения для API (уменьшение размера)"""
    try:
        image = Image.open(io.BytesIO(image_bytes))
        
        # Конвертируем в RGB если нужно
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
            image = background
        
        # Уменьшаем размер если слишком большой
        if max(image.size) > max_size:
            ratio = max_size / max(image.size)
            new_size = tuple(int(dim * ratio) for dim in image.size)
            image = image.resize(new_size, Image.Resampling.LANCZOS)
        
        # Сохраняем с оптимизацией
        output = io.BytesIO()
        image.save(output, format='JPEG', quality=85, optimize=True)
        
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Ошибка оптимизации изображения: {e}")
        return image_bytes  # Возвращаем как есть в случае ошибки
    
@dp.message_handler(content_types=ContentType.PHOTO, state=AskState.waiting_for_question)
async def handle_photo_in_question_mode(message: Message, state: FSMContext):
    """Обработка фото в режиме вопросов"""
    user_id = message.from_user.id
    
    if not check_access_api(user_id):
        await message.answer("⛔ *Доступ закрыт.*", reply_markup=pay_kb)
        await state.finish()
        return
    
    # Отправляем сообщение о начале обработки
    wait_msg = await message.answer("🖼️ Анализирую изображение...")
    
    try:
        # Получаем файл изображения
        photo = message.photo[-1]  # Берем самое большое изображение
        file_info = await bot.get_file(photo.file_id)
        downloaded_file = await bot.download_file(file_info.file_path)
        image_bytes = downloaded_file.read()
        
        # Оптимизируем изображение для API
        optimized_bytes = await optimize_image_for_api(image_bytes)
        
        # Получаем текст запроса (подпись к фото)
        user_text = message.caption or "Проанализируй это изображение с маркетинговой точки зрения."
        
        # Обрабатываем изображение
        result = await process_image_with_gpt4v(optimized_bytes, user_text, user_id)
        
        await wait_msg.delete()  # Удаляем сообщение "Анализирую..."
        
        # Обрабатываем результат
        if result.get("type") == "text":
            answer = result.get("content", "")
            hist = user_history.setdefault(user_id, [])
            hist.append({"role": "user", "content": f"[Изображение] {user_text}"})
            hist.append({"role": "assistant", "content": answer})
            hist[:] = hist[-10:]  # Ограничиваем историю
            
            formatted_answer = format_answer(answer)
            await message.answer(formatted_answer)
            
        elif result.get("type") == "excel":
            await process_excel_response(message, result)
            
        else:
            await message.answer(f"📋 *Результат анализа:*\n\n{result.get('content', 'Нет данных')}")
            
    except Exception as e:
        logger.error(f"Ошибка обработки изображения: {e}")
        await wait_msg.edit_text("⚠️ *Ошибка обработки изображения.* Попробуйте отправить другое изображение или текстовый запрос.")

@dp.message_handler(state=AskState.waiting_for_question)
async def process_question(message: Message, state: FSMContext):
    user_id = message.from_user.id

    if not check_access_api(user_id):
        await message.answer("⛔ *Доступ закрыт.* Для получения доступа, пожалуйста, оплатите подписку. 📞 Свяжитесь с менеджером для активации.", reply_markup=pay_kb)
        return await state.finish()

    hist = user_history.setdefault(user_id, [])
    hist.append({"role": "user", "content": message.text})
    hist[:] = hist[-10:]

    messages = [system_prompt] + hist

    try:
        completion = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages
        )
        raw = completion.choices[0].message.content
        data = smart_json_fix(raw)
    except Exception as e:
        logger.error(f"AI JSON error: {e}")
        await message.answer("⚠️ *Ошибка обработки запроса.* Попробуйте немного позже.")
        return

    if data["type"] == "text":
        await process_text_response(message, data, hist)
    elif data["type"] == "excel":
        await process_excel_response(message, data)


async def process_text_response(message: Message, data: dict, hist: list):
    answer = data["content"]
    hist.append({"role": "assistant", "content": answer})
    formatted_answer = format_answer(answer)
    await message.answer(formatted_answer)


async def process_excel_response(message: Message, data: dict):
    prompt = data["prompt"]
    try:
        completion_excel = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Всегда создавай таблицу минимум из 10 строк. "
                        "Каждая строка должна быть уникальной, разнообразной и детально заполненной. "
                        "Возвращай STRICT JSON строго в формате: "
                        "{\"headers\": [...], \"rows\": [...]}. "
                        "Где headers — список названий колонок, а rows — массив массивов строк. "
                        "Никаких других ключей, полей или текста. "
                        "Без описаний. "
                        "Только валидный JSON."
                    )
                },
                {"role": "user", "content": f"Создай таблицу с подробным анализом для стратегии продвижения: {prompt}"}
            ]
        )
        raw_table = completion_excel.choices[0].message.content
        logger.info(f"Ответ от GPT: {raw_table}")
        cleaned_response = raw_table.strip()
        try:
            table_json = json.loads(cleaned_response)
        except json.JSONDecodeError as e:
            logger.error(f"Ошибка при парсинге JSON: {e}")
            await message.answer("⚠️ *Ошибка обработки данных.* Попробуйте немного позже или обратитесь к менеджеру.")
            return

        if not validate_table_structure(table_json):
            raise ValueError(f"Структура таблицы неверная: {table_json}")
        await create_detailed_excel(message, table_json)
    except Exception as e:
        logger.error(f"Ошибка создания Excel: {e}")
        await message.answer("⚠️ *Ошибка создания Excel файла.* Попробуйте ещё раз или обратитесь к менеджеру.")

async def create_detailed_excel(message: Message, table_json: dict):
    headers = table_json["headers"]
    rows = table_json["rows"]
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Основные задачи"
    ws1.append(headers)
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in rows:
        ws1.append(row)
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws1.column_dimensions[column].width = max_length + 4
    ws2 = wb.create_sheet("Ответственные и сроки")
    ws2.append(["Пункт анализа", "Ответственный", "Срок", "Статус"])

    for row in rows:
        ws2.append([
            row[0],           # Пункт анализа
            "Не начато"
        ]) 
    ws3 = wb.create_sheet("Статусы задач")
    ws3.append(["Пункт анализа", "Статус", "Прогресс"])

    for row in rows:
        ws3.append([
            row[0],
            "Не начата"
        ])
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        filepath = tmp.name
    await message.answer_document(
        InputFile(filepath),
        caption="📊 *Подробный Excel файл с данными для стратегии продвижения готов!*"
    )

def split_json_objects(response: str) -> list:
    return response.split("\n}\n{")

def validate_table_structure(table_json: dict) -> bool:
    return isinstance(table_json, dict) and "headers" in table_json and "rows" in table_json

async def create_and_send_excel(message: Message, table_json: dict):
    headers = table_json["headers"]
    rows = table_json["rows"]
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws.append(headers)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        filepath = tmp.name
    await message.answer_document(
        InputFile(filepath),
        caption="📊 *Ваш Excel файл готов!*"
    )

def format_answer(answer: str) -> str:
    formatted = answer
    formatted = formatted.replace("\n", "\n\n")  
    formatted = re.sub(r'(\d+\.)', r'### \1', formatted) 
    formatted = formatted.replace("Социальные сети", "📱 **Социальные сети**")
    return formatted

@dp.message_handler(content_types=types.ContentType.VOICE, state="*")
async def ignore_voice(message: Message, state: FSMContext):
    await message.answer("🎤 Голосовые сообщения пока не поддерживаются. Напишите текстом ✍️")

@dp.message_handler()
async def fallback(message: Message):
    await message.answer("Выберите действие в меню ⬇️", reply_markup=main_menu)

@dp.message_handler(Command("grant"))
async def manager_grant(message: Message):
    if message.from_user.id != MANAGER_ID:
        return await message.answer("⛔ Нет прав.")
    tg_id = int(message.get_args().strip())
    if activate_user_api(tg_id):
        await message.answer("✅ Доступ активирован.")
        try:
            await bot.send_message(tg_id, "Ваш доступ активирован!")
        except:
            pass
    else:
        await message.answer("❌ Ошибка активации.")

if __name__ == "__main__":
    from aiogram import executor
    logger.info("Бот запущен...")
    executor.start_polling(dp, skip_updates=True)